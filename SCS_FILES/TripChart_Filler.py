import pandas as pd
import csv
import re
from openpyxl import load_workbook
from datetime import datetime, timedelta
import sys
import os
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("member_id", type=str, help="Member ID.")
parser.add_argument("line_no", type=str, help="Line No.")
args = parser.parse_args()
member_id = args.member_id
line_no = args.line_no

new_location = f"ALL_USER_TT/LINE{line_no}_{member_id}"
blank_tc_path = os.path.dirname(os.path.abspath(__file__)) + f"/{new_location}/USEFUL OUTPUT_{member_id}/SOLUTION 1/BLANK TC {member_id} LINE {line_no}.xlsx"
excel_path = os.path.dirname(os.path.abspath(__file__)) + f"/{new_location}/{member_id}.xlsx"


path = os.path.dirname(os.path.abspath(__file__)) + f"/{new_location}/USEFUL OUTPUT_{member_id}/SOLUTION 1/"
# Read CSV file with variable number of columns per row
with open(f'{path}TrainLoopVerification_Line{line_no}{member_id}.CSV', newline='') as f:
    reader = csv.reader(f)
    data = list(reader)

# Find the maximum number of columns in any row
# max_cols = max(len(row) for row in data)

# # Pad rows with fewer columns with empty strings
# padded_data = [row + [''] * (max_cols - len(row)) for row in data]

# Create DataFrame
csv_df = pd.DataFrame(data)

# Import Excel file (first sheet by default)
excel_df = pd.read_excel(f'{path}BLANK TC {member_id} LINE {line_no}.xlsx')

# print("CSV File Content:")
# print(csv_df)

# print("\nExcel File Content:")
# print(excel_df)



# Function to truncate time strings to hh:mm
def truncate_to_hhmm(val):
    if isinstance(val, str):
        # Try to match time patterns
        match = re.match(r'(\d{1,2}):(\d{2})', val)
        if match:
            return f"{int(match.group(1)):02d}:{match.group(2)}"
    elif pd.api.types.is_datetime64_any_dtype(type(val)) or hasattr(val, 'hour'):
        return f"{val.hour:02d}:{val.minute:02d}"
    return val

# Apply to all cells in csv_df
def normalize_time(val):
    if isinstance(val, str):
        match = re.match(r'(\d{1,2}):(\d{2})', val)
        if match:
            hour = int(match.group(1))
            minute = int(match.group(2))
            hour = hour % 24
            return f"{hour:02d}:{minute:02d}"
    return val

def parse_time_str(time_str):
    try:
        return datetime.strptime(time_str, "%H:%M")
    except Exception:
        print(f"Error parsing time string: {time_str}")
        return None

def is_time_within(val, ref_time, delta_minutes=10):
    t1 = parse_time_str(val)
    t2 = parse_time_str(ref_time)
    if t1 and t2:
        diff = abs((t1 - t2).total_seconds()) / 60
        return diff <= delta_minutes
    return False


csv_df = csv_df.map(normalize_time)
csv_df = csv_df.map(truncate_to_hhmm)

# Apply to all cells in excel_df
# excel_df = excel_df.applymap(format_mixed_duration)
excel_df = excel_df.map(truncate_to_hhmm)

print("\nCSV File Content (Truncated to hh:mm):")
print(csv_df)

print("\nExcel File Content (Truncated to hh:mm):")
print(excel_df)


# print(csv_df.iloc[0,2])
# print(excel_df.iloc[21,2])
# if excel_df.iloc[21,2] == csv_df.iloc[0,2]:
#     print("The values match!")

# Columns to process in excel_df
if line_no == 'AEL':
    columns_to_check = [3, 5, 7, 9, 12, 14]
elif line_no == '34':
    columns_to_check = [3, 5, 7, 15, 17, 20, 24]

for row_idx in range(len(excel_df)):
    for col in columns_to_check:
        if col >= len(excel_df.columns):
            continue  # Skip if column index is out of bounds

        # Get train number from first column of the row
        train_no = excel_df.iloc[row_idx, 0]
        # print(train_no)

        # Skip if train_no is empty
        # if not isinstance(train_no, str) or not train_no.strip():
        #     continue

        # Find triptime1 (left of col)
        triptime1 = None
        triptime1_col = None
        for c in range(col - 1, -1, -1):
            val = excel_df.iloc[row_idx, c]
            # print(val)
            if isinstance(val, str) and re.match(r'\d{1,2}:\d{2}', val):
                triptime1 = str(val)
                triptime1_col = c
                break

        # Find triptime2 (right of col)
        triptime2 = None
        triptime2_col = None
        for c in range(col + 1, len(excel_df.columns)):
            val = excel_df.iloc[row_idx, c]
            if isinstance(val, str) and re.match(r'\d{1,2}:\d{2}', val):
                triptime2 = str(val)
                triptime2_col = c
                break

        if not triptime1 or not triptime2:
            continue  # Skip if both triptime are not found
        # print(triptime1, triptime2)
        # Find the row in csv_df where first column matches train_no
        csv_row_idx = None
        count = 0
        for idx in range(len(csv_df)):
            if str(csv_df.iloc[idx, 0]).strip() == str(train_no).strip():
                csv_row_idx = idx
                triptime1_csv_col = None
                triptime2_csv_col = None
                
                val_1 = str(csv_df.iloc[csv_row_idx, 2])
                if val_1 == triptime1:
                    triptime1_csv_col = 2
                elif is_time_within(triptime1, val_1):
                    triptime1_csv_col = 2

                val_2 = str(csv_df.iloc[csv_row_idx, 4])
                if val_2 == triptime2:
                    triptime2_csv_col = 4
                elif is_time_within(triptime2, val_2):
                    triptime2_csv_col = 4

                if triptime1_csv_col is None or triptime2_csv_col is None:
                    continue

                # Ensure triptime1_csv_col < triptime2_csv_col
                # left = min(triptime1_csv_col, triptime2_csv_col)
                # right = max(triptime1_csv_col, triptime2_csv_col)

                # print(left, right)
                # break
                # print(type(triptime1))
                # print(type(triptime2))
                # print(type(val_1))
                # print(type(val_2))
                # sys.exit(0)
                # Search for integer value (duty no.) between triptime1 and triptime2 in csv_df row
                dutyno = None
                # if left == right:
                #     if triptime1_csv_col == None:
                #         for c in range(right,right-5,-1):
                #             val = csv_df.iloc[csv_row_idx, c]
                #             if isinstance(val, str) and val.isdigit():
                #                 dutyno = val
                #                 break
                #     elif triptime2_csv_col == None:
                #         for c in range(left, left + 5):
                #             val = csv_df.iloc[csv_row_idx, c]
                #             if isinstance(val, str) and val.isdigit():
                #                 dutyno = val
                #                 break
                
                
                val = csv_df.iloc[csv_row_idx, 5]
                if isinstance(val, str) and val.isdigit():
                    dutyno = val
                
                    
                # print(dutyno)
                # break

                # If found, append to excel_df at col+1
                if dutyno:
                    target_col = col
                    if target_col < len(excel_df.columns):
                        excel_df.iloc[row_idx, target_col] = int(dutyno)
                        break

# Optionally, save the updated excel_df
# print(excel_df)
excel_df.to_excel(f'{path}FILLED TC {member_id} LINE {line_no}.xlsx', index=False)
print("Trip chart updated and saved to 'DIWALI SPCL TRIP CHART_FILLED.xlsx'.")
# Copy columns D,F,G,O,Q,T,Y from FILLED to BLANK without changing format

# Column letters to copy
if line_no == 'AEL':
    columns_to_copy = ['D', 'F', 'H', 'J', 'M', 'O']
elif line_no == '34':
    columns_to_copy = ['D', 'F', 'H', 'P', 'R', 'U', 'Y']

# Load both workbooks
wb_filled = load_workbook(f'{path}FILLED TC {member_id} LINE {line_no}.xlsx')
wb_blank = load_workbook(f'{path}BLANK TC {member_id} LINE {line_no}.xlsx')

ws_filled = wb_filled.active
ws_blank = wb_blank.active

for col_letter in columns_to_copy:
    for row in range(3, ws_filled.max_row + 1):
        cell_filled = ws_filled[f"{col_letter}{row}"]
        cell_blank = ws_blank[f"{col_letter}{row}"]
        cell_blank.value = cell_filled.value

wb_blank.save(f'{path}TC {member_id} LINE {line_no}.xlsx.xlsx')
print("Selected columns copied from FILLED to BLANK file.")